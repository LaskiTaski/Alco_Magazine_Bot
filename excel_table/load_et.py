from openpyxl import load_workbook
from data_base import sql_db_other


wb = load_workbook(filename ='КрасноеБелое.xlsx')
ws = wb['Magazine']


def read_file():
    for i in range(2,100):
        line = []
        for j in 'ABCDE':
            line.append(ws[f'{j}{i}'].value)

        if None not in line:
            sql_db_other.sql_add(line)
